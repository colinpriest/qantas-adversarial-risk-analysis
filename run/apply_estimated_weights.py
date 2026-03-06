"""
Apply estimated Board utility weights to governance_spec.xlsx.

Reads parameter_estimates.csv from the quantification pipeline and updates
the utilities_board sheet in governance_spec.xlsx.  The original file is
backed up before modification.

Usage:
    python -m run.apply_estimated_weights outputs/parameter_estimates.csv
    python -m run.apply_estimated_weights outputs/parameter_estimates.csv --dry-run
    python -m run.apply_estimated_weights outputs/parameter_estimates.csv --spec data/governance_spec.xlsx
"""

import argparse
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# Mapping from quantification parameter names to engine parameter names.
# Direct 1:1 mappings.
DIRECT_MAP = {
    "w1": "early_ceo_departure_cost",
    "w2": "vote_penalty_weight",
    "w3": "overwhelming_penalty_weight",
    "w4": "spill_risk_weight",
    "w8s": "ceo_loss_shock_strike",
    "w8o": "ceo_loss_shock_overwhelming",
    "w8r": "ceo_loss_shock_adverse",
    "w9": "reputational_spill_weight",
    "w12": "board_d1_liability",
    "w13": "qantas_legal_d1_penalty",
    "w15": "adverse_review_ceo_present_penalty",
}

# Collapsed parameter decompositions.
# Each entry: quantification_name → {engine_name: spec_default_proportion}
COLLAPSED_MAP = {
    "w_removal": {
        "implementation_cost_sack": 0.3,
        "ceo_loss_cost": 1.5,
    },
    "w_inaction": {
        "second_strike_spill_penalty": 8.0,
        "board_regulatory_liability": 5.0,
        "qantas_legal_d_rev_penalty": 2.0,
    },
}


def load_estimates(csv_path: Path) -> dict[str, float]:
    """Load estimated parameter values from quantification output CSV."""
    df = pd.read_csv(csv_path)
    required_cols = {"parameter", "estimate"}
    if not required_cols.issubset(df.columns):
        raise ValueError(f"CSV must have columns: {required_cols}. Found: {set(df.columns)}")
    return dict(zip(df["parameter"], df["estimate"]))


def decompose_to_engine_params(estimates: dict[str, float]) -> dict[str, float]:
    """Convert quantification parameter estimates to engine parameter names.

    Direct parameters are mapped 1:1.  Collapsed parameters are decomposed
    proportionally to their spec default ratios.
    """
    engine_params = {}

    # Direct mappings
    for q_name, engine_name in DIRECT_MAP.items():
        if q_name in estimates:
            engine_params[engine_name] = float(estimates[q_name])

    # Collapsed decompositions
    for q_name, constituents in COLLAPSED_MAP.items():
        if q_name in estimates:
            total = float(estimates[q_name])
            spec_total = sum(constituents.values())
            if spec_total < 1e-12:
                # Distribute equally if all spec defaults are zero
                for engine_name in constituents:
                    engine_params[engine_name] = total / len(constituents)
            else:
                for engine_name, spec_default in constituents.items():
                    engine_params[engine_name] = total * spec_default / spec_total

    return engine_params


def update_governance_spec(
    spec_path: Path,
    engine_params: dict[str, float],
    dry_run: bool = False,
) -> list[dict]:
    """Update the utilities_board sheet in governance_spec.xlsx.

    Returns a list of changes made (for logging/preview).
    """
    wb = load_workbook(spec_path)
    if "utilities_board" not in wb.sheetnames:
        raise ValueError(f"Sheet 'utilities_board' not found in {spec_path}")

    ws = wb["utilities_board"]

    # Find the parameter_name and value columns
    header_row = 1
    param_col = None
    value_col = None
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col_idx).value
        if cell_val == "parameter_name":
            param_col = col_idx
        elif cell_val == "value":
            value_col = col_idx

    if param_col is None or value_col is None:
        raise ValueError("Could not find 'parameter_name' and 'value' columns "
                         f"in utilities_board sheet. Headers: "
                         f"{[ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]}")

    changes = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        param_name = ws.cell(row=row_idx, column=param_col).value
        if param_name in engine_params:
            old_val = ws.cell(row=row_idx, column=value_col).value
            new_val = round(engine_params[param_name], 6)
            changes.append({
                "parameter": param_name,
                "old_value": old_val,
                "new_value": new_val,
                "row": row_idx,
            })
            if not dry_run:
                ws.cell(row=row_idx, column=value_col).value = new_val

    if not dry_run and changes:
        wb.save(spec_path)

    wb.close()
    return changes


def main():
    parser = argparse.ArgumentParser(
        description="Apply estimated Board utility weights to governance_spec.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "This script reads parameter estimates from the quantification pipeline\n"
            "and updates the utilities_board sheet in governance_spec.xlsx.\n\n"
            "Collapsed parameters (w_removal, w_inaction) are decomposed proportionally\n"
            "to their spec default ratios.\n\n"
            "A timestamped backup of governance_spec.xlsx is created before any changes.\n\n"
            "Examples:\n"
            "  python -m run.apply_estimated_weights outputs/parameter_estimates.csv\n"
            "  python -m run.apply_estimated_weights outputs/parameter_estimates.csv --dry-run\n"
        ),
    )
    parser.add_argument("estimates_csv", type=str,
                        help="Path to parameter_estimates.csv from quantification pipeline")
    parser.add_argument("--spec", type=str,
                        default=str(PROJECT_ROOT / "data" / "governance_spec.xlsx"),
                        help="Path to governance_spec.xlsx (default: data/governance_spec.xlsx)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show changes without writing to file")

    args = parser.parse_args()

    csv_path = Path(args.estimates_csv)
    spec_path = Path(args.spec)

    if not csv_path.exists():
        logger.error(f"Estimates file not found: {csv_path}")
        sys.exit(1)
    if not spec_path.exists():
        logger.error(f"Governance spec not found: {spec_path}")
        sys.exit(1)

    # Load estimates
    logger.info(f"Loading estimates from {csv_path}")
    estimates = load_estimates(csv_path)
    logger.info(f"  Found {len(estimates)} parameter estimates")

    # Decompose to engine parameters
    engine_params = decompose_to_engine_params(estimates)
    logger.info(f"  Decomposed to {len(engine_params)} engine parameters")

    # Show decomposition
    for name, value in sorted(engine_params.items()):
        logger.info(f"    {name} = {value:.6f}")

    if args.dry_run:
        logger.info("DRY RUN — showing changes without writing")
    else:
        # Backup
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = spec_path.with_name(f"governance_spec_backup_{timestamp}.xlsx")
        shutil.copy2(spec_path, backup_path)
        logger.info(f"Backup created: {backup_path}")

    # Apply changes
    changes = update_governance_spec(spec_path, engine_params, dry_run=args.dry_run)

    if not changes:
        logger.warning("No matching parameters found in utilities_board sheet!")
        sys.exit(1)

    # Report
    logger.info(f"\n{'Parameter':<40} {'Old':>10} {'New':>10} {'Change':>10}")
    logger.info("-" * 72)
    for c in changes:
        old = c["old_value"]
        new = c["new_value"]
        old_str = f"{old:.4f}" if isinstance(old, (int, float)) else str(old)
        change = new - old if isinstance(old, (int, float)) else "N/A"
        change_str = f"{change:+.4f}" if isinstance(change, (int, float)) else change
        logger.info(f"  {c['parameter']:<38} {old_str:>10} {new:.4f!s:>10} {change_str:>10}")

    if args.dry_run:
        logger.info(f"\nDry run complete. {len(changes)} parameters would be updated.")
        logger.info("Run without --dry-run to apply changes.")
    else:
        logger.info(f"\nUpdated {len(changes)} parameters in {spec_path}")
        logger.info("The engine will use these values on next run:")
        logger.info("  python -m run.run_board_mode --checkpoint C0 --n_draws 100")


if __name__ == "__main__":
    main()
